#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
import argparse
import configdb
from consolemsg import error, step, success, warn
from openpyxl import load_workbook

#uri = 'https://api.somenergia.coop/form/contractacio'
#uri='http://testing.somenergia.coop:5001/form/contractacio'

def add_contract(data):
    import requests
    r = requests.post(uri, data=data)
    return (r.status_code, r.reason, r.text)

def crea_contractes(filename = None):
    if not filename:
        sys.exit("No hi ha cap arxiu per crear contractes")
    wb = load_workbook(filename)
    ws = wb.active
    valors = []
    keys = []
    pols = []
    column = 1
    while ws.cell(row=1,column=column).value:
        keys.append(ws.cell(row=1,column=column).value)
        column += 1
    row=2
    while ws.cell(row=row,column=1).value:
        column = 1
        valors = []
        while ws.cell(row=1,column=column).value:
            valors.append(ws.cell(row=row,column=column).value)
            column += 1
        pols.append(dict(zip(keys,valors)))
        row+=1 
    for pol in pols:
        succes(pol['cups'])
        status,reason, text = add_contract(pol)
        step('Status: ' + str(status))
        step('Reason: ' + str(reason))
        step('Text: ' + str(text))

        
def main(csv_file, check_conn=True):
    uri = configdb.api_contractacio['uri']

    if check_conn:
        msg = "You are requesting to: {}, do you want to continue? (Y/n)"
        step(msg.format(uri))
        answer = raw_input()
        while answer.lower() not in ['y', 'n', '']:
            answer = raw_input()
            step("Do you want to continue? (Y/n)")

        if answer in ['n', 'N']:
            raise KeyboardInterrupt
        else:
            try:
                crea_contractes(csv_file)
            except Exception as e:
                warn(e)    

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Erstellen Sie Verträge in loser Schüttung'
    )

    parser.add_argument(
        '--file',
        dest='csv_file',
        required=True,
        help="csv amb les noves fitxes cliente a crear"
    )

    parser.add_argument(
        '--check-conn',
        type=bool,
        nargs='?',
        default=False,
        const=True,
        help="Check para comprobar que URL queremos atacar"
    )

    args = parser.parse_args()
    print("Check conn: ", args.check_conn)
    try:
        main(args.csv_file, args.check_conn)
    except (KeyboardInterrupt, SystemExit, SystemError):
        warn("Aarrggghh you kill me :(")
    except IndexError:
        warn("No se interprear lo que me dices, sorry :'(")
    else:
        success("Chao!")
